import pandas as pd
import numpy as np
import unicodedata

def encontrar_aba_correta(xls_file, nome_alvo):
    def normalizar(txt):
        return unicodedata.normalize('NFKD', str(txt)).encode('ASCII', 'ignore').decode('utf-8').lower().strip()
    
    nome_alvo_norm = normalizar(nome_alvo)
    for sheet in xls_file.sheet_names:
        if normalizar(sheet) == nome_alvo_norm:
            return sheet
    return None

def ler_range_exato(path, sheet_name, range_str, usecols=None):
    try:
        xls = pd.ExcelFile(path)
        aba = encontrar_aba_correta(xls, sheet_name)
        if not aba: return None

        start_row, end_row = map(int, range_str.split(':'))
        skiprows = start_row - 1
        nrows = (end_row - start_row) + 1

        df = pd.read_excel(
            xls, 
            sheet_name=aba, 
            header=None, 
            skiprows=skiprows, 
            nrows=nrows,
            usecols=usecols
        )
        
        # Garante índices numéricos (0, 1, 2...)
        if not df.empty:
            df.columns = range(df.shape[1])
            
        return df
    except Exception as e:
        print(f"⚠️ Erro leitura {range_str}: {e}")
        return None

# --- PROCESSADOR ESPECÍFICO PARA EVOLUÇÃO ANUAL (K3:N15) ---
def processar_evolucao_anual_anos(df):
    """
    Lê a tabela onde as colunas são ANOS (2023, 2024, 2025).
    Linha 0: Cabeçalho (Mês, 2023, 2024, 2025)
    Linhas 1+: Dados (Jan, Fev...)
    """
    if df is None or df.empty: return {}

    # 1. Pega os anos do cabeçalho
    header = df.iloc[0]
    colunas_anos = {} # {indice: "2023"}
    
    # Começa da coluna 1 porque a 0 é "Mês"
    for i in range(1, len(header)):
        ano_label = str(header[i]).replace(".0", "").strip() # Remove .0 se vier float
        colunas_anos[i] = ano_label

    # 2. Processa os dados
    df_dados = df.iloc[1:]
    labels = df_dados.iloc[:, 0].astype(str).tolist() # Meses
    
    valores = {}
    for idx, ano in colunas_anos.items():
        valores[ano] = pd.to_numeric(df_dados.iloc[:, idx], errors='coerce').fillna(0).tolist()

    return {"labels": labels, "valores": valores}

# --- PROCESSADOR DE REGIONAL (NOVO FORMATO: Meses VS Regionais) ---
def processar_pendencias_regional_meses(df):
    """
    Para Regional (K20:O32)
    Linha 0: Mês, BRA, RSOU, SAO1, SAO2
    Linhas 1+: janeiro, fevereiro, etc.
    Retorna X-axis = regionais, Datasets = meses
    """
    if df is None or df.empty: return {}

    header = df.iloc[0]
    regionais = []
    
    for i in range(1, len(header)):
        reg_label = str(header[i]).strip() 
        regionais.append(reg_label)

    df_dados = df.iloc[1:]
    
    valores_por_mes = {}
    
    for _, row in df_dados.iterrows():
        mes = str(row.iloc[0]).strip()
        if mes.lower() in ["nan", "none", ""]: continue
        
        valores_mes = []
        for i in range(1, len(header)):
            val = pd.to_numeric(row.iloc[i], errors='coerce')
            val = 0 if pd.isna(val) else int(val)
            valores_mes.append(val)
            
        valores_por_mes[mes] = valores_mes

    return {"labels": regionais, "valores": valores_por_mes}

# --- PROCESSADOR DE REGIONAL Antigo (Mantido) ---
def processar_regional_ok_nok(df):
    """
    Mantido para backward compatibility se precisar.
    """
    if df is None or df.empty: return {}
    idx_total, idx_ok, idx_nok = 1, -2, -1
    header = df.iloc[0]
    for i, val in enumerate(header):
        t = str(val).upper()
        if "TOTAL" in t: idx_total = i
        if "OK" in t or "CONFORME" in t: idx_ok = i
        if "NOK" in t or "IRREGULAR" in t: idx_nok = i
        
    df_dados = df.iloc[1:]
    labels = df_dados.iloc[:, 0].astype(str).tolist()
    valores = {}
    valores["Total"] = pd.to_numeric(df_dados.iloc[:, idx_total], errors='coerce').fillna(0).tolist()
    valores["OK"] = pd.to_numeric(df_dados.iloc[:, idx_ok], errors='coerce').fillna(0).tolist()
    valores["NOK"] = pd.to_numeric(df_dados.iloc[:, idx_nok], errors='coerce').fillna(0).tolist()
    return {"labels": labels, "valores": valores}

# --- PROCESSADOR DE STATUS (Back Room / Gelo) ---
def processar_status_bloco(df):
    if df is None or df.empty: return {"valores": {}, "labels": []}
    
    header = df.iloc[0]
    idx_total = -1; idx_ok = -1; idx_nok = -1; idx_pend = -1

    for i, val in enumerate(header):
        texto = str(val).upper()
        if "TOTAL" in texto or "PROGRAMADO" in texto: idx_total = i
        if "CONFORME" in texto or "OK" in texto or "SATISFAT" in texto: idx_ok = i
        if "IRREGULAR" in texto or "NOK" in texto or "INSATISFAT" in texto: idx_nok = i
        if "PENDENTE" in texto: idx_pend = i

    if idx_total == -1: idx_total = 1
    if idx_ok == -1: idx_ok = df.shape[1] - 3
    if idx_nok == -1: idx_nok = df.shape[1] - 4
    if idx_pend == -1: idx_pend = df.shape[1] - 1

    df_dados = df.iloc[1:]
    labels = []; lista_total = []; lista_ok = []; lista_nok = []; lista_pend = []

    for i, row in df_dados.iterrows():
        sigla = str(row.iloc[0]).strip()
        if "TOTAL" in sigla.upper(): continue
        try:
            vt = int(pd.to_numeric(row.iloc[idx_total], errors='coerce') or 0) if idx_total != -1 else 0
            vo = int(pd.to_numeric(row.iloc[idx_ok], errors='coerce') or 0) if idx_ok != -1 else 0
            vn = int(pd.to_numeric(row.iloc[idx_nok], errors='coerce') or 0) if idx_nok != -1 else 0
            vp = int(pd.to_numeric(row.iloc[idx_pend], errors='coerce') or 0) if idx_pend != -1 else 0
            
            labels.append(sigla)
            lista_total.append(vt)
            lista_ok.append(vo)
            lista_nok.append(vn)
            lista_pend.append(vp)
        except: continue

    return {"labels": labels, "valores": {"Total": lista_total, "OK": lista_ok, "NOK": lista_nok, "Pendentes": lista_pend}}

# --- PROCESSADOR DE PENDÊNCIAS DE GELO (MATRIZ) ---
def processar_pendencias_top(df):
    if df is None or df.empty: return {"valores": {}, "labels": []}

    # Linha 0 = cabeçalho: ['Sigla', 'Máquina de Gelo', 'Bin da torre', 'Bin Mc Café']
    header = [str(x).strip() for x in df.iloc[0]]
    print(f"🔍 [DEBUG PEND GELO] Header: {header}")

    # Descobre as colunas de dados (tudo que não é "Sigla" ou vazio/nan)
    categorias = []
    col_indices = []
    for i, nome in enumerate(header):
        if nome and nome.lower() not in ["sigla", "nan", "", "none"]:
            categorias.append(nome)
            col_indices.append(i)

    print(f"🔍 [DEBUG PEND GELO] Categorias: {categorias}, Indices: {col_indices}")

    # Linhas de dados (pula o cabeçalho)
    df_dados = df.iloc[1:]
    labels_regionais = []
    valores_por_categoria = {cat: [] for cat in categorias}

    for _, row in df_dados.iterrows():
        sigla = str(row.iloc[0]).strip()
        if not sigla or sigla.lower() in ["nan", "none", ""]:
            continue
        labels_regionais.append(sigla)
        for cat, idx in zip(categorias, col_indices):
            try:
                val = pd.to_numeric(row.iloc[idx], errors='coerce')
                val = 0 if pd.isna(val) else int(val)
            except:
                val = 0
            valores_por_categoria[cat].append(val)

    print(f"🔍 [DEBUG PEND GELO] Regionais: {labels_regionais}")
    for cat in categorias:
        print(f"   {cat}: {valores_por_categoria[cat]}")

    return {"labels": labels_regionais, "valores": valores_por_categoria}

# --- DETALHES GERAIS (Mantido) ---
def processar_detalhes_parametros_por_mes(df):
    IDX_MES = 3 
    regras_grupos = { "Back Room": slice(9, 18), "Gelo Pool": slice(19, 22), "Máquina de Gelo": slice(23, 26), "Bin Café": slice(27, 30), "Bin Bebidas": slice(31, 34) }
    ORDEM_MESES = ["janeiro", "fevereiro", "março", "abril", "maio", "junho", "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"]
    try:
        if df.shape[1] <= IDX_MES: return {}
        
        # Filtra apenas 2026 pela coluna de data (coluna index ~4, procura coluna com 'data')
        col_data_idx = None
        for i, col_name in enumerate(df.columns):
            if 'data' in str(col_name).lower():
                col_data_idx = i
                break
        
        if col_data_idx is not None:
            df_work = df.copy()
            df_work.iloc[:, col_data_idx] = pd.to_datetime(df_work.iloc[:, col_data_idx], errors='coerce')
            mask_2026 = df_work.iloc[:, col_data_idx].dt.year == 2026
            df = df[mask_2026]
        
        col_mes = df.iloc[:, IDX_MES].astype(str).str.strip().str.lower().str.capitalize()
        resultado_grupos = {}
        for nome_grupo, fatiador in regras_grupos.items():
            df_grupo = df.iloc[:, fatiador].copy()
            lista_graficos = []
            for col in df_grupo.columns:
                param_nome = str(col).strip()
                if "unnamed" in param_nome.lower(): continue
                df_temp = pd.DataFrame({'Mes': col_mes, 'Valor': df_grupo[col].astype(str).str.lower().str.strip()})
                df_temp = df_temp[df_temp['Valor'].isin(['ok', 'nok'])]
                if df_temp.empty:
                    lista_graficos.append({"titulo": param_nome, "labels": [], "ok": [], "nok": []})
                    continue
                agrupado = df_temp.groupby(['Mes', 'Valor']).size().unstack(fill_value=0)
                if 'ok' not in agrupado.columns: agrupado['ok'] = 0
                if 'nok' not in agrupado.columns: agrupado['nok'] = 0
                meses_presentes = agrupado.index.tolist()
                meses_ordenados = sorted(meses_presentes, key=lambda x: ORDEM_MESES.index(x.lower()) if x.lower() in ORDEM_MESES else 999)
                agrupado = agrupado.reindex(meses_ordenados)
                lista_graficos.append({"titulo": param_nome, "labels": agrupado.index.tolist(), "ok": agrupado['ok'].tolist(), "nok": agrupado['nok'].tolist()})
            resultado_grupos[nome_grupo] = lista_graficos
        return resultado_grupos
    except: return {}

def processar_backroom_mensal(caminho_arquivo):
    """
    Lê a coluna I (Back Room) da aba GERAL e agrupa por mês (coluna D).
    - OK (case-insensitive): conforme
    - NA (case-insensitive): desconsiderar (excluir da contagem)
    - Qualquer outro valor: NOK (não conforme)
    Retorna: { labels: [meses], ok: [qtd], nok: [qtd], ok_pct: [%], nok_pct: [%] }
    """
    ORDEM_MESES = [
        "janeiro", "fevereiro", "março", "abril", "maio", "junho",
        "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"
    ]
    
    try:
        import openpyxl
        wb = openpyxl.load_workbook(caminho_arquivo, read_only=True, data_only=True)
        ws = wb['GERAL']
        
        # Coletar dados: mês, data, backroom
        registros = []
        for row in ws.iter_rows(min_row=3, min_col=1, max_col=9):
            mes_val = row[3].value   # Col D - Mês
            data_val = row[4].value  # Col E - Data (para filtrar ano)
            br_val = row[8].value    # Col I - Back Room
            
            if br_val is None or mes_val is None:
                continue
            
            br_str = str(br_val).strip().lower()
            mes_str = str(mes_val).strip().lower()
            
            # Ignorar NA
            if br_str == 'na':
                continue
            
            # Filtrar somente 2026
            ano = None
            if data_val is not None:
                try:
                    if hasattr(data_val, 'year'):
                        ano = data_val.year
                    else:
                        from datetime import datetime
                        dt = pd.to_datetime(data_val, errors='coerce')
                        if pd.notna(dt):
                            ano = dt.year
                except:
                    pass
            
            if ano is not None and ano != 2026:
                continue
            
            # Classificar OK/NOK
            is_ok = br_str == 'ok'
            registros.append({'mes': mes_str, 'is_ok': is_ok})
        
        wb.close()
        
        if not registros:
            return {"labels": [], "ok": [], "nok": [], "ok_pct": [], "nok_pct": []}
        
        # Agrupar por mês
        from collections import defaultdict
        agrupado = defaultdict(lambda: {'ok': 0, 'nok': 0})
        
        for reg in registros:
            mes = reg['mes']
            if reg['is_ok']:
                agrupado[mes]['ok'] += 1
            else:
                agrupado[mes]['nok'] += 1
        
        # Ordenar meses cronologicamente
        meses_presentes = list(agrupado.keys())
        meses_ordenados = sorted(meses_presentes, key=lambda x: ORDEM_MESES.index(x) if x in ORDEM_MESES else 999)
        
        labels = []
        ok_vals = []
        nok_vals = []
        ok_pcts = []
        nok_pcts = []
        
        for mes in meses_ordenados:
            dados = agrupado[mes]
            total = dados['ok'] + dados['nok']
            labels.append(mes.capitalize())
            ok_vals.append(dados['ok'])
            nok_vals.append(dados['nok'])
            ok_pcts.append(round((dados['ok'] / total) * 100, 1) if total > 0 else 0)
            nok_pcts.append(round((dados['nok'] / total) * 100, 1) if total > 0 else 0)
        
        print(f"✅ [BACKROOM MENSAL] {len(registros)} registros, {len(labels)} meses")
        for i, m in enumerate(labels):
            print(f"   {m}: OK={ok_vals[i]} ({ok_pcts[i]}%) | NOK={nok_vals[i]} ({nok_pcts[i]}%)")
        
        return {
            "labels": labels,
            "ok": ok_vals,
            "nok": nok_vals,
            "ok_pct": ok_pcts,
            "nok_pct": nok_pcts
        }
    
    except Exception as e:
        print(f"❌ [BACKROOM MENSAL] Erro: {e}")
        import traceback
        traceback.print_exc()
        return {"labels": [], "ok": [], "nok": [], "ok_pct": [], "nok_pct": []}


def processar_backroom_por_regional(caminho_arquivo):
    """
    Lê a coluna I (Back Room) da aba GERAL e agrupa por Regional (dinâmico) e Mês (coluna D).
    Retorna: { 
        meses: ["Janeiro", ...], 
        regionais: ["RSOU", "BRA", "SAO1", "SAO2"],
        dados: {
            "Janeiro": { "RSOU": {ok: N, nok: N, ok_pct: N, nok_pct: N}, ... },
            ...
        }
    }
    """
    ORDEM_MESES = [
        "janeiro", "fevereiro", "março", "abril", "maio", "junho",
        "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"
    ]
    
    try:
        import openpyxl
        from collections import defaultdict
        
        wb = openpyxl.load_workbook(caminho_arquivo, read_only=True, data_only=True)
        ws = wb['GERAL']
        
        # Descobrir qual coluna é "Regional" lendo a linha de cabeçalho (linha 2)
        header_row = None
        for row in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=20):
            header_row = row
            break
        
        regional_col_idx = None
        if header_row:
            for i, cell in enumerate(header_row):
                val = str(cell.value or '').strip().lower()
                if 'regional' in val or 'regiao' in val:
                    regional_col_idx = i
                    break
        
        if regional_col_idx is None:
            # Fallback: tentar coluna B (índice 1)
            regional_col_idx = 1
            print(f"⚠️ [BACKROOM REGIONAL] Coluna 'Regional' não encontrada no header, usando coluna B (idx={regional_col_idx})")
        else:
            print(f"✅ [BACKROOM REGIONAL] Coluna 'Regional' encontrada no índice {regional_col_idx}")
        
        # Coletar dados: regional, mês, data, backroom
        registros = []
        max_col_needed = max(regional_col_idx, 8) + 1  # Pelo menos até coluna I (idx 8)
        
        for row in ws.iter_rows(min_row=3, min_col=1, max_col=max_col_needed + 1):
            if len(row) <= max(regional_col_idx, 8):
                continue
                
            regional_val = row[regional_col_idx].value
            mes_val = row[3].value      # Col D - Mês
            data_val = row[4].value     # Col E - Data
            br_val = row[8].value       # Col I - Back Room
            
            if br_val is None or mes_val is None or regional_val is None:
                continue
            
            br_str = str(br_val).strip().lower()
            mes_str = str(mes_val).strip().lower()
            regional_str = str(regional_val).strip().upper()
            
            # Ignorar NA e vazios
            if br_str == 'na' or not regional_str or regional_str in ['NAN', 'NONE', '', '#N/A', '#N/D', 'NA']:
                continue
            
            # Filtrar somente 2026
            ano = None
            if data_val is not None:
                try:
                    if hasattr(data_val, 'year'):
                        ano = data_val.year
                    else:
                        dt = pd.to_datetime(data_val, errors='coerce')
                        if pd.notna(dt):
                            ano = dt.year
                except:
                    pass
            
            if ano is not None and ano != 2026:
                continue
            
            is_ok = br_str == 'ok'
            registros.append({
                'regional': regional_str, 
                'mes': mes_str, 
                'is_ok': is_ok
            })
        
        wb.close()
        
        if not registros:
            return {"meses": [], "regionais": [], "dados": {}}
        
        # Agrupar por mês e regional
        agrupado = defaultdict(lambda: defaultdict(lambda: {'ok': 0, 'nok': 0}))
        todas_regionais = set()
        
        for reg in registros:
            agrupado[reg['mes']][reg['regional']]['ok' if reg['is_ok'] else 'nok'] += 1
            todas_regionais.add(reg['regional'])
        
        # Ordenar meses e regionais
        meses_presentes = list(agrupado.keys())
        meses_ordenados = sorted(meses_presentes, key=lambda x: ORDEM_MESES.index(x) if x in ORDEM_MESES else 999)
        regionais_ordenadas = sorted(list(todas_regionais))
        
        # Construir resposta
        dados = {}
        for mes in meses_ordenados:
            mes_cap = mes.capitalize()
            dados[mes_cap] = {}
            for regional in regionais_ordenadas:
                info = agrupado[mes][regional]
                total = info['ok'] + info['nok']
                dados[mes_cap][regional] = {
                    'ok': info['ok'],
                    'nok': info['nok'],
                    'ok_pct': round((info['ok'] / total) * 100, 1) if total > 0 else 0,
                    'nok_pct': round((info['nok'] / total) * 100, 1) if total > 0 else 0,
                }
        
        print(f"✅ [BACKROOM REGIONAL] {len(registros)} registros, {len(meses_ordenados)} meses, {len(regionais_ordenadas)} regionais")
        for mes in meses_ordenados:
            mes_cap = mes.capitalize()
            for reg in regionais_ordenadas:
                d = dados[mes_cap][reg]
                print(f"   {mes_cap}/{reg}: OK={d['ok']} ({d['ok_pct']}%) | NOK={d['nok']} ({d['nok_pct']}%)")
        
        return {
            "meses": [m.capitalize() for m in meses_ordenados],
            "regionais": regionais_ordenadas,
            "dados": dados
        }
    
    except Exception as e:
        print(f"❌ [BACKROOM REGIONAL] Erro: {e}")
        import traceback
        traceback.print_exc()
        return {"meses": [], "regionais": [], "dados": {}}


def processar_gelopool_mensal(caminho_arquivo):
    """
    Lê a coluna S (Gelo Pool, index 18) da aba GERAL e agrupa por mês (coluna D).
    Mesma lógica de processar_backroom_mensal mas para Gelo Pool.
    - OK (case-insensitive): conforme
    - NA (case-insensitive): desconsiderar (excluir da contagem)
    - Qualquer outro valor: NOK (não conforme)
    Retorna: { labels: [meses], ok: [qtd], nok: [qtd], ok_pct: [%], nok_pct: [%] }
    """
    ORDEM_MESES = [
        "janeiro", "fevereiro", "março", "abril", "maio", "junho",
        "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"
    ]
    
    try:
        import openpyxl
        wb = openpyxl.load_workbook(caminho_arquivo, read_only=True, data_only=True)
        ws = wb['GERAL']
        
        registros = []
        for row in ws.iter_rows(min_row=3, min_col=1, max_col=19):  # Até coluna S (idx 18)
            mes_val = row[3].value   # Col D - Mês
            data_val = row[4].value  # Col E - Data (para filtrar ano)
            gp_val = row[18].value   # Col S - Gelo (pool)
            
            if gp_val is None or mes_val is None:
                continue
            
            gp_str = str(gp_val).strip().lower()
            mes_str = str(mes_val).strip().lower()
            
            # Ignorar NA
            if gp_str == 'na':
                continue
            
            # Filtrar somente 2026
            ano = None
            if data_val is not None:
                try:
                    if hasattr(data_val, 'year'):
                        ano = data_val.year
                    else:
                        from datetime import datetime
                        dt = pd.to_datetime(data_val, errors='coerce')
                        if pd.notna(dt):
                            ano = dt.year
                except:
                    pass
            
            if ano is not None and ano != 2026:
                continue
            
            # Classificar OK/NOK
            is_ok = gp_str == 'ok'
            registros.append({'mes': mes_str, 'is_ok': is_ok})
        
        wb.close()
        
        if not registros:
            return {"labels": [], "ok": [], "nok": [], "ok_pct": [], "nok_pct": []}
        
        # Agrupar por mês
        from collections import defaultdict
        agrupado = defaultdict(lambda: {'ok': 0, 'nok': 0})
        
        for reg in registros:
            mes = reg['mes']
            if reg['is_ok']:
                agrupado[mes]['ok'] += 1
            else:
                agrupado[mes]['nok'] += 1
        
        # Ordenar meses cronologicamente
        meses_presentes = list(agrupado.keys())
        meses_ordenados = sorted(meses_presentes, key=lambda x: ORDEM_MESES.index(x) if x in ORDEM_MESES else 999)
        
        labels = []
        ok_vals = []
        nok_vals = []
        ok_pcts = []
        nok_pcts = []
        
        for mes in meses_ordenados:
            dados = agrupado[mes]
            total = dados['ok'] + dados['nok']
            labels.append(mes.capitalize())
            ok_vals.append(dados['ok'])
            nok_vals.append(dados['nok'])
            ok_pcts.append(round((dados['ok'] / total) * 100, 1) if total > 0 else 0)
            nok_pcts.append(round((dados['nok'] / total) * 100, 1) if total > 0 else 0)
        
        print(f"✅ [GELO POOL MENSAL] {len(registros)} registros, {len(labels)} meses")
        for i, m in enumerate(labels):
            print(f"   {m}: OK={ok_vals[i]} ({ok_pcts[i]}%) | NOK={nok_vals[i]} ({nok_pcts[i]}%)")
        
        return {
            "labels": labels,
            "ok": ok_vals,
            "nok": nok_vals,
            "ok_pct": ok_pcts,
            "nok_pct": nok_pcts
        }
    
    except Exception as e:
        print(f"❌ [GELO POOL MENSAL] Erro: {e}")
        import traceback
        traceback.print_exc()
        return {"labels": [], "ok": [], "nok": [], "ok_pct": [], "nok_pct": []}


def processar_gelopool_por_regional(caminho_arquivo):
    """
    Lê a coluna S (Gelo Pool, index 18) da aba GERAL e agrupa por Regional e Mês (coluna D).
    Mesma lógica de processar_backroom_por_regional mas para Gelo Pool.
    Retorna: { 
        meses: ["Janeiro", ...], 
        regionais: ["RSOU", "BRA", "SAO1", "SAO2"],
        dados: {
            "Janeiro": { "RSOU": {ok: N, nok: N, ok_pct: N, nok_pct: N}, ... },
            ...
        }
    }
    """
    ORDEM_MESES = [
        "janeiro", "fevereiro", "março", "abril", "maio", "junho",
        "julho", "agosto", "setembro", "outubro", "novembro", "dezembro"
    ]
    
    try:
        import openpyxl
        from collections import defaultdict
        
        wb = openpyxl.load_workbook(caminho_arquivo, read_only=True, data_only=True)
        ws = wb['GERAL']
        
        # Descobrir qual coluna é "Regional" lendo a linha de cabeçalho (linha 2)
        header_row = None
        for row in ws.iter_rows(min_row=2, max_row=2, min_col=1, max_col=20):
            header_row = row
            break
        
        regional_col_idx = None
        if header_row:
            for i, cell in enumerate(header_row):
                val = str(cell.value or '').strip().lower()
                if 'regional' in val or 'regiao' in val:
                    regional_col_idx = i
                    break
        
        if regional_col_idx is None:
            regional_col_idx = 2  # Coluna C (Regional)
            print(f"⚠️ [GELO POOL REGIONAL] Coluna 'Regional' não encontrada no header, usando coluna C (idx={regional_col_idx})")
        else:
            print(f"✅ [GELO POOL REGIONAL] Coluna 'Regional' encontrada no índice {regional_col_idx}")
        
        # Coletar dados: regional, mês, data, gelo pool
        registros = []
        max_col_needed = max(regional_col_idx, 18) + 1  # Pelo menos até coluna S (idx 18)
        
        for row in ws.iter_rows(min_row=3, min_col=1, max_col=max_col_needed + 1):
            if len(row) <= max(regional_col_idx, 18):
                continue
                
            regional_val = row[regional_col_idx].value
            mes_val = row[3].value      # Col D - Mês
            data_val = row[4].value     # Col E - Data
            gp_val = row[18].value      # Col S - Gelo (pool)
            
            if gp_val is None or mes_val is None or regional_val is None:
                continue
            
            gp_str = str(gp_val).strip().lower()
            mes_str = str(mes_val).strip().lower()
            regional_str = str(regional_val).strip().upper()
            
            # Ignorar NA e vazios
            if gp_str == 'na' or not regional_str or regional_str in ['NAN', 'NONE', '', '#N/A', '#N/D', 'NA']:
                continue
            
            # Filtrar somente 2026
            ano = None
            if data_val is not None:
                try:
                    if hasattr(data_val, 'year'):
                        ano = data_val.year
                    else:
                        dt = pd.to_datetime(data_val, errors='coerce')
                        if pd.notna(dt):
                            ano = dt.year
                except:
                    pass
            
            if ano is not None and ano != 2026:
                continue
            
            is_ok = gp_str == 'ok'
            registros.append({
                'regional': regional_str, 
                'mes': mes_str, 
                'is_ok': is_ok
            })
        
        wb.close()
        
        if not registros:
            return {"meses": [], "regionais": [], "dados": {}}
        
        # Agrupar por mês e regional
        agrupado = defaultdict(lambda: defaultdict(lambda: {'ok': 0, 'nok': 0}))
        todas_regionais = set()
        
        for reg in registros:
            agrupado[reg['mes']][reg['regional']]['ok' if reg['is_ok'] else 'nok'] += 1
            todas_regionais.add(reg['regional'])
        
        # Ordenar meses e regionais
        meses_presentes = list(agrupado.keys())
        meses_ordenados = sorted(meses_presentes, key=lambda x: ORDEM_MESES.index(x) if x in ORDEM_MESES else 999)
        regionais_ordenadas = sorted(list(todas_regionais))
        
        # Construir resposta
        dados = {}
        for mes in meses_ordenados:
            mes_cap = mes.capitalize()
            dados[mes_cap] = {}
            for regional in regionais_ordenadas:
                info = agrupado[mes][regional]
                total = info['ok'] + info['nok']
                dados[mes_cap][regional] = {
                    'ok': info['ok'],
                    'nok': info['nok'],
                    'ok_pct': round((info['ok'] / total) * 100, 1) if total > 0 else 0,
                    'nok_pct': round((info['nok'] / total) * 100, 1) if total > 0 else 0,
                }
        
        print(f"✅ [GELO POOL REGIONAL] {len(registros)} registros, {len(meses_ordenados)} meses, {len(regionais_ordenadas)} regionais")
        for mes in meses_ordenados:
            mes_cap = mes.capitalize()
            for reg in regionais_ordenadas:
                d = dados[mes_cap][reg]
                print(f"   {mes_cap}/{reg}: OK={d['ok']} ({d['ok_pct']}%) | NOK={d['nok']} ({d['nok_pct']}%)")
        
        return {
            "meses": [m.capitalize() for m in meses_ordenados],
            "regionais": regionais_ordenadas,
            "dados": dados
        }
    
    except Exception as e:
        print(f"❌ [GELO POOL REGIONAL] Erro: {e}")
        import traceback
        traceback.print_exc()
        return {"meses": [], "regionais": [], "dados": {}}


def processar_aba_geral(caminho_arquivo):
    try:
        try: df_header = pd.read_excel(caminho_arquivo, sheet_name='GERAL', header=1)
        except: df_header = pd.read_csv(caminho_arquivo, header=1)
        detalhes = processar_detalhes_parametros_por_mes(df_header)
        return {"detalhes_parametros": detalhes}, None
    except Exception as e: return None, str(e)